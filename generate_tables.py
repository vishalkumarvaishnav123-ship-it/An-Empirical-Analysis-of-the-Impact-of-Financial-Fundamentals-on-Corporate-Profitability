"""
Thesis: An Empirical Analysis of the Impact of Financial Fundamentals on Corporate Profitability
Author: Vishal
Description: This script performs all calculations and generates all 10 tables used in the thesis.
Requirements: pip install pandas numpy statsmodels openpyxl
"""

import pandas as pd
import numpy as np
import statsmodels.api as sm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# ============================================
# CONFIGURATION
# ============================================
DATA_FILE = 'German_compnies_data_corrected_with_uk_and_germany___Kopie_00.xlsx'
OUTPUT_DIR = 'tables'
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Sheet mapping
SHEET_MAP = {
    'Germany 2021': 'german2021',
    'Germany 2024': 'german 2024',
    'UK 2021': 'UK2021',
    'UK 2024': 'UK2024'
}

DATA_LABELS = ['Germany 2021', 'Germany 2024', 'UK 2021', 'UK 2024']
INDEP_VARS = ['Equity_TA', 'Quick_Ratio', 'Sales_TA']
DEP_VARS = ['ROA', 'ROE']
ALL_VARS = ['Equity_TA', 'Quick_Ratio', 'Sales_TA', 'ROA', 'ROE']

# Styling
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
HEADER_FILL = PatternFill('solid', fgColor='4472C4')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
DATA_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(bold=True, name='Arial', size=10)
TITLE_FONT = Font(bold=True, name='Arial', size=12)
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
GREY_FILL = PatternFill('solid', fgColor='D9D9D9')


# ============================================
# HELPER FUNCTIONS
# ============================================
def style_header(ws, row, cols):
    """Apply header styling to a row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_data(ws, row, cols):
    """Apply data styling to a row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def sig_stars(p):
    """Return significance stars based on p-value."""
    if p <= 0.01:
        return '***'
    elif p <= 0.05:
        return '**'
    elif p <= 0.10:
        return '*'
    return ''


def sig_color(p):
    """Return font color based on significance level."""
    if p <= 0.01:
        return Font(bold=True, name='Arial', size=10, color='FF0000')
    elif p <= 0.05:
        return Font(bold=True, name='Arial', size=10, color='0000FF')
    elif p <= 0.10:
        return Font(bold=True, name='Arial', size=10, color='008000')
    return DATA_FONT


# ============================================
# STEP 1: DATA LOADING AND CLEANING
# ============================================
def load_clean(sheet_name):
    """
    Load and clean a single sheet from the Excel file.

    Steps:
    1. Read the raw Excel sheet
    2. Map columns to standardised names (handles typos like 'Euity_TA')
    3. Keep only the 6 required columns
    4. Convert all numeric columns to float
    5. Drop rows with missing values
    """
    df = pd.read_excel(DATA_FILE, sheet_name=sheet_name)

    # Map columns to standard names
    col_map = {}
    for c in df.columns:
        cl = c.strip().lower()
        if 'firm' in cl:
            col_map[c] = 'Firm'
        elif 'euity' in cl or 'equity_ta' in cl:
            col_map[c] = 'Equity_TA'
        elif 'quick' in cl:
            col_map[c] = 'Quick_Ratio'
        elif 'sal' in cl and 'ta' in cl:
            col_map[c] = 'Sales_TA'
        elif 'roa' in cl:
            col_map[c] = 'ROA'
        elif 'roe' in cl:
            col_map[c] = 'ROE'

    df = df.rename(columns=col_map)
    needed = ['Firm', 'Equity_TA', 'Quick_Ratio', 'Sales_TA', 'ROA', 'ROE']
    df = df[needed].dropna(subset=['Firm'])

    for c in needed[1:]:
        df[c] = pd.to_numeric(df[c], errors='coerce')

    df = df.dropna()
    return df.reset_index(drop=True)


def load_all_data():
    """Load all four datasets."""
    datasets = {}
    for label, sheet in SHEET_MAP.items():
        datasets[label] = load_clean(sheet)
        print(f"  Loaded {label}: {len(datasets[label])} firms")
    return datasets


# ============================================
# STEP 2: DESCRIPTIVE STATISTICS
# ============================================
def calculate_descriptive_stats(datasets):
    """
    Calculate descriptive statistics for all variables across all 4 samples.

    For each sample, computes:
    - Count, Mean, Std Dev, Min, 25th percentile, Median, 75th percentile, Max
    """
    desc_stats = {}
    for label in DATA_LABELS:
        df = datasets[label][ALL_VARS]
        desc = df.describe().round(4)
        desc_stats[label] = desc
        print(f"\n  {label}:")
        print(f"    Mean ROA = {df['ROA'].mean():.4f}")
        print(f"    Mean ROE = {df['ROE'].mean():.4f}")
        print(f"    Mean Equity/TA = {df['Equity_TA'].mean():.4f}")
        print(f"    Mean Quick Ratio = {df['Quick_Ratio'].mean():.4f}")
        print(f"    Mean Sales/TA = {df['Sales_TA'].mean():.4f}")
    return desc_stats


# ============================================
# STEP 3: CORRELATION ANALYSIS
# ============================================
def calculate_correlations(datasets):
    """
    Calculate Pearson correlation matrices for all 4 samples.

    Purpose:
    1. Check pairwise relationships between variables
    2. Check for multicollinearity (correlations > 0.80 among IVs = concern)
    """
    corr_data = {}
    for label in DATA_LABELS:
        corr = datasets[label][ALL_VARS].corr().round(4)
        corr_data[label] = corr

        # Check multicollinearity
        max_iv_corr = 0
        for i, v1 in enumerate(INDEP_VARS):
            for v2 in INDEP_VARS[i + 1:]:
                c = abs(corr.loc[v1, v2])
                if c > max_iv_corr:
                    max_iv_corr = c
        print(f"  {label}: Max IV correlation = {max_iv_corr:.4f} "
              f"{'(OK)' if max_iv_corr < 0.80 else '(WARNING: multicollinearity)'}")

    return corr_data


# ============================================
# STEP 4: OLS REGRESSION (TWO-STEP APPROACH)
# ============================================
def run_regressions(datasets):
    """
    Run OLS regressions for all 8 models using the two-step approach.

    Step 1: Full model with all 3 independent variables
        Y_i = alpha + beta1 * Equity_TA_i + beta2 * Quick_Ratio_i + beta3 * Sales_TA_i + epsilon_i

    Step 2: If any variable has p > 0.10, drop it and re-run with significant variables only.

    For each model, stores:
    - Coefficients, standard errors, t-statistics, p-values
    - R-squared, Adjusted R-squared
    - F-statistic, F p-value
    - Sample size
    """
    step1_results = {}
    step2_results = {}

    for label in DATA_LABELS:
        df = datasets[label]
        for dep in DEP_VARS:
            key = f"{label} - {dep}"

            # --- STEP 1: Full Model ---
            X = sm.add_constant(df[INDEP_VARS])
            y = df[dep]
            model = sm.OLS(y, X).fit()
            step1_results[key] = model

            print(f"\n  {key} (Step 1):")
            print(f"    R² = {model.rsquared:.4f}, "
                  f"Adj R² = {model.rsquared_adj:.4f}, "
                  f"F = {model.fvalue:.4f}, "
                  f"F p-value = {model.f_pvalue:.4f}")
            for v in model.params.index:
                stars = sig_stars(model.pvalues[v])
                print(f"    {v:15s}: coef = {model.params[v]:8.4f}, "
                      f"se = {model.bse[v]:8.4f}, "
                      f"t = {model.tvalues[v]:8.4f}, "
                      f"p = {model.pvalues[v]:.4f} {stars}")

            # --- STEP 2: Reduced Model ---
            sig_vars = [v for v in INDEP_VARS if model.pvalues[v] <= 0.10]

            if 0 < len(sig_vars) < 3:
                X2 = sm.add_constant(df[sig_vars])
                model2 = sm.OLS(y, X2).fit()
                step2_results[key] = model2
                print(f"    Step 2: Retained {sig_vars}")
                print(f"    R² = {model2.rsquared:.4f}, "
                      f"Adj R² = {model2.rsquared_adj:.4f}")
                for v in model2.params.index:
                    stars = sig_stars(model2.pvalues[v])
                    print(f"      {v:15s}: coef = {model2.params[v]:8.4f}, "
                          f"p = {model2.pvalues[v]:.4f} {stars}")
            elif len(sig_vars) == 3:
                step2_results[key] = model
                print(f"    Step 2: All variables significant — full model retained.")
            else:
                step2_results[key] = None
                print(f"    Step 2: No significant variables — no reduced model.")

    return step1_results, step2_results


# ============================================
# STEP 5: HYPOTHESIS TESTING
# ============================================
def evaluate_hypotheses(step1_results):
    """
    Evaluate each hypothesis against the regression results.

    Hypotheses (all expect positive sign):
        H1a: Equity/TA -> ROA (+)
        H1b: Quick Ratio -> ROA (+)
        H1c: Sales/TA -> ROA (+)
        H2a: Equity/TA -> ROE (+)
        H2b: Quick Ratio -> ROE (+)
        H2c: Sales/TA -> ROE (+)

    Classification:
        Supported     = significant (p <= 0.10) AND correct sign (+)
        Rejected      = significant (p <= 0.10) AND wrong sign (-)
        Not Significant = not significant (p > 0.10)
    """
    hypotheses = [
        ('H1a', 'ROA', 'Equity_TA', 'Equity/TA', '+'),
        ('H1b', 'ROA', 'Quick_Ratio', 'Quick Ratio', '+'),
        ('H1c', 'ROA', 'Sales_TA', 'Sales/TA', '+'),
        ('H2a', 'ROE', 'Equity_TA', 'Equity/TA', '+'),
        ('H2b', 'ROE', 'Quick_Ratio', 'Quick Ratio', '+'),
        ('H2c', 'ROE', 'Sales_TA', 'Sales/TA', '+'),
    ]

    results = []
    for hyp_id, dep, var, var_label, expected in hypotheses:
        row = {'Hypothesis': hyp_id, 'Dep': dep, 'Variable': var_label,
               'Expected': expected}
        for label in DATA_LABELS:
            key = f"{label} - {dep}"
            model = step1_results[key]
            coef = model.params[var]
            pval = model.pvalues[var]
            sig = pval <= 0.10
            actual = '+' if coef > 0 else '-'
            if sig and actual == expected:
                outcome = 'Supported'
            elif sig and actual != expected:
                outcome = 'Rejected'
            else:
                outcome = 'Not Significant'
            row[label] = outcome
        results.append(row)

    # Print summary
    supported = sum(1 for r in results for l in DATA_LABELS if r[l] == 'Supported')
    rejected = sum(1 for r in results for l in DATA_LABELS if r[l] == 'Rejected')
    not_sig = sum(1 for r in results for l in DATA_LABELS if r[l] == 'Not Significant')
    total = supported + rejected + not_sig
    print(f"\n  Supported:       {supported}/{total} ({supported/total*100:.1f}%)")
    print(f"  Rejected:        {rejected}/{total} ({rejected/total*100:.1f}%)")
    print(f"  Not Significant: {not_sig}/{total} ({not_sig/total*100:.1f}%)")

    return results


# ============================================
# STEP 6: GENERATE EXCEL TABLES
# ============================================
def generate_table_2_1(wb):
    """Table 2.1: Summary of Key Empirical Studies."""
    ws = wb.active
    ws.title = 'Table 2.1'
    ws.cell(1, 1, 'Table 2.1: Summary of Key Empirical Studies').font = TITLE_FONT

    headers = ['Author(s)', 'Year', 'Country', 'Sample', 'Dependent Variable',
               'Key Independent Variables', 'Key Findings']
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, 7)

    studies = [
        ['Abor (2005)', '2005', 'Ghana', '22 firms', 'ROE',
         'Debt/Equity, Short-term Debt',
         'Positive relationship between short-term debt and ROE'],
        ['Gill et al. (2011)', '2011', 'USA', '272 firms', 'ROE',
         'Debt/Equity, Debt/Assets',
         'Positive effect of debt ratio on profitability'],
        ['Salim & Yadav (2012)', '2012', 'Malaysia', '237 firms', 'ROA, ROE',
         'Debt/Assets, Equity/Assets',
         'Negative effect of leverage on ROA and ROE'],
        ['Mehari & Aemiro (2013)', '2013', 'Ethiopia', '9 insurers', 'ROA',
         'Leverage, Liquidity, Size',
         'Liquidity positively affects ROA'],
        ['Tailab (2014)', '2014', 'USA', '30 firms', 'ROA, ROE',
         'Debt/Equity, Debt/Assets',
         'No significant relationship found'],
        ['Ahmad et al. (2015)', '2015', 'Pakistan', '100 firms', 'ROA',
         'Current Ratio, Quick Ratio',
         'Positive effect of liquidity on ROA'],
        ['Alarussi & Alhaderi (2018)', '2018', 'Malaysia', '120 firms', 'ROA, ROE',
         'Liquidity, Firm Size',
         'Positive effect of equity ratio on profitability'],
        ['Nguyen & Nguyen (2020)', '2020', 'Vietnam', '30 firms', 'ROA, ROE',
         'Debt/Assets, Sales Growth',
         'Negative leverage effect; positive asset turnover effect'],
    ]
    for r, study in enumerate(studies, 4):
        for c, val in enumerate(study, 1):
            ws.cell(r, c, val)
        style_data(ws, r, 7)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 50


def generate_table_2_2(wb):
    """Table 2.2: Summary of Hypotheses and Expected Signs."""
    ws = wb.create_sheet('Table 2.2')
    ws.cell(1, 1, 'Table 2.2: Summary of Hypotheses and Expected Signs').font = TITLE_FONT

    headers = ['Hypothesis', 'Dependent Variable', 'Independent Variable',
               'Expected Sign', 'Rationale']
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, 5)

    hyps = [
        ['H1a', 'ROA', 'Equity/Total Assets', '+',
         'Higher equity ratio indicates lower financial risk and stable earnings'],
        ['H1b', 'ROA', 'Quick Ratio', '+',
         'Higher liquidity enables firms to meet obligations and invest efficiently'],
        ['H1c', 'ROA', 'Sales/Total Assets', '+',
         'Higher asset turnover indicates more efficient use of assets'],
        ['H2a', 'ROE', 'Equity/Total Assets', '+',
         'Higher equity financing reduces interest burden, increasing net income'],
        ['H2b', 'ROE', 'Quick Ratio', '+',
         'Adequate liquidity supports operations and enhances returns to shareholders'],
        ['H2c', 'ROE', 'Sales/Total Assets', '+',
         'Efficient asset utilisation generates higher revenues and shareholder returns'],
    ]
    for r, h in enumerate(hyps, 4):
        for c, val in enumerate(h, 1):
            ws.cell(r, c, val)
        style_data(ws, r, 5)
        ws.cell(r, 4).font = Font(bold=True, name='Arial', size=10, color='008000')

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 60


def generate_table_3_1(wb, datasets):
    """Table 3.1: Sample Companies (Germany and UK)."""
    ws = wb.create_sheet('Table 3.1')
    ws.cell(1, 1, 'Table 3.1: Sample Companies (Germany and UK)').font = TITLE_FONT

    headers = ['No.', 'German Companies', 'No.', 'UK Companies']
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, 4)

    ger_firms = datasets['Germany 2021']['Firm'].tolist()
    uk_firms = datasets['UK 2021']['Firm'].tolist()
    for r in range(max(len(ger_firms), len(uk_firms))):
        ws.cell(r + 4, 1, r + 1)
        ws.cell(r + 4, 2, ger_firms[r] if r < len(ger_firms) else '')
        ws.cell(r + 4, 3, r + 1)
        ws.cell(r + 4, 4, uk_firms[r] if r < len(uk_firms) else '')
        style_data(ws, r + 4, 4)

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 35


def generate_table_3_2(wb):
    """Table 3.2: Variable Definitions and Formulas."""
    ws = wb.create_sheet('Table 3.2')
    ws.cell(1, 1, 'Table 3.2: Variable Definitions and Formulas').font = TITLE_FONT

    headers = ['Variable', 'Type', 'Abbreviation', 'Formula', 'Description']
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, 5)

    var_defs = [
        ['Return on Assets', 'Dependent', 'ROA',
         'Net Profit / Total Assets',
         'Measures how efficiently a firm uses its assets to generate profit'],
        ['Return on Equity', 'Dependent', 'ROE',
         'Net Profit / Total Equity',
         'Measures profit generated per unit of shareholders equity'],
        ['Equity to Total Assets', 'Independent', 'Equity/TA',
         'Total Equity / Total Assets',
         'Indicates the proportion of assets financed by equity'],
        ['Quick Ratio', 'Independent', 'QR',
         '(Current Assets - Inventory) / Current Liabilities',
         'Measures short-term liquidity excluding inventory'],
        ['Sales to Total Assets', 'Independent', 'Sales/TA',
         'Sales / Total Assets',
         'Measures how efficiently assets generate revenue (asset turnover)'],
    ]
    for r, v in enumerate(var_defs, 4):
        for c, val in enumerate(v, 1):
            ws.cell(r, c, val)
        style_data(ws, r, 5)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 55


def generate_table_3_3(wb, desc_stats):
    """Table 3.3: Descriptive Statistics (All Countries and Years)."""
    ws = wb.create_sheet('Table 3.3')
    ws.cell(1, 1, 'Table 3.3: Descriptive Statistics (All Countries and Years)').font = TITLE_FONT

    row = 3
    for label in DATA_LABELS:
        ws.cell(row, 1, label).font = Font(
            bold=True, name='Arial', size=11, color='4472C4'
        )
        row += 1
        headers = ['Statistic', 'Equity/TA', 'Quick Ratio', 'Sales/TA', 'ROA', 'ROE']
        for i, h in enumerate(headers, 1):
            ws.cell(row, i, h)
        style_header(ws, row, 6)
        row += 1

        desc = desc_stats[label]
        for stat in desc.index:
            ws.cell(row, 1, stat)
            ws.cell(row, 1).font = BOLD_FONT
            for c, col in enumerate(desc.columns, 2):
                ws.cell(row, c, round(desc.loc[stat, col], 4))
            style_data(ws, row, 6)
            row += 1
        row += 1

    ws.column_dimensions['A'].width = 15
    for c in 'BCDEF':
        ws.column_dimensions[c].width = 15


def generate_table_4_1(wb, corr_data):
    """Table 4.1: Correlation Matrices (Germany and UK)."""
    ws = wb.create_sheet('Table 4.1')
    ws.cell(1, 1, 'Table 4.1: Correlation Matrices (Germany and UK)').font = TITLE_FONT

    var_display = {
        'Equity_TA': 'Equity/TA',
        'Quick_Ratio': 'Quick Ratio',
        'Sales_TA': 'Sales/TA',
        'ROA': 'ROA',
        'ROE': 'ROE'
    }

    row = 3
    for label in DATA_LABELS:
        ws.cell(row, 1, label).font = Font(
            bold=True, name='Arial', size=11, color='4472C4'
        )
        row += 1
        headers = [''] + [var_display[v] for v in ALL_VARS]
        for i, h in enumerate(headers, 1):
            ws.cell(row, i, h)
        style_header(ws, row, 6)
        row += 1

        corr = corr_data[label]
        for var in ALL_VARS:
            ws.cell(row, 1, var_display[var])
            ws.cell(row, 1).font = BOLD_FONT
            for c, col in enumerate(ALL_VARS, 2):
                ws.cell(row, c, round(corr.loc[var, col], 4))
            style_data(ws, row, 6)
            row += 1
        row += 1

    ws.column_dimensions['A'].width = 15
    for c in 'BCDEF':
        ws.column_dimensions[c].width = 15


def generate_table_4_2(wb, step1_results):
    """Table 4.2: Regression Results – ROA (All 4 Models)."""
    ws = wb.create_sheet('Table 4.2')
    ws.cell(1, 1, 'Table 4.2: Regression Results – ROA (All 4 Models)').font = TITLE_FONT

    # Headers row 1
    headers1 = ['Variable', 'Germany 2021', '', 'Germany 2024', '',
                'UK 2021', '', 'UK 2024', '']
    for i, h in enumerate(headers1, 1):
        ws.cell(3, i, h)
    # Headers row 2
    headers2 = ['', 'Coeff.', 'p-value', 'Coeff.', 'p-value',
                'Coeff.', 'p-value', 'Coeff.', 'p-value']
    for i, h in enumerate(headers2, 1):
        ws.cell(4, i, h)
    style_header(ws, 3, 9)
    style_header(ws, 4, 9)

    vars_order = ['const', 'Equity_TA', 'Quick_Ratio', 'Sales_TA']
    var_display = ['Intercept', 'Equity/TA', 'Quick Ratio', 'Sales/TA']

    row = 5
    for var, vd in zip(vars_order, var_display):
        ws.cell(row, 1, vd)
        ws.cell(row, 1).font = BOLD_FONT
        col = 2
        for dl in DATA_LABELS:
            key = f"{dl} - ROA"
            model = step1_results[key]
            coef = round(model.params[var], 4)
            pval = round(model.pvalues[var], 4)
            ws.cell(row, col, coef)
            ws.cell(row, col + 1, pval)
            ws.cell(row, col).font = sig_color(pval)
            col += 2
        style_data(ws, row, 9)
        row += 1

    # Model statistics
    row += 1
    for stat_name, stat_attr in [('R²', 'rsquared'), ('Adjusted R²', 'rsquared_adj'),
                                  ('F-statistic', 'fvalue'), ('F p-value', 'f_pvalue'),
                                  ('N', 'nobs')]:
        ws.cell(row, 1, stat_name)
        ws.cell(row, 1).font = BOLD_FONT
        col = 2
        for dl in DATA_LABELS:
            key = f"{dl} - ROA"
            model = step1_results[key]
            val = getattr(model, stat_attr)
            ws.cell(row, col, round(float(val), 4) if stat_attr != 'nobs' else int(val))
            col += 2
        style_data(ws, row, 9)
        row += 1

    row += 1
    ws.cell(row, 1,
            'Significance: Red = p<0.01, Blue = p<0.05, Green = p<0.10').font = Font(
        italic=True, name='Arial', size=9
    )

    ws.column_dimensions['A'].width = 15
    for c in 'BCDEFGHI':
        ws.column_dimensions[c].width = 14


def generate_table_4_3(wb, step1_results):
    """Table 4.3: Regression Results – ROE (All 4 Models)."""
    ws = wb.create_sheet('Table 4.3')
    ws.cell(1, 1, 'Table 4.3: Regression Results – ROE (All 4 Models)').font = TITLE_FONT

    headers1 = ['Variable', 'Germany 2021', '', 'Germany 2024', '',
                'UK 2021', '', 'UK 2024', '']
    headers2 = ['', 'Coeff.', 'p-value', 'Coeff.', 'p-value',
                'Coeff.', 'p-value', 'Coeff.', 'p-value']
    for i, h in enumerate(headers1, 1):
        ws.cell(3, i, h)
    for i, h in enumerate(headers2, 1):
        ws.cell(4, i, h)
    style_header(ws, 3, 9)
    style_header(ws, 4, 9)

    vars_order = ['const', 'Equity_TA', 'Quick_Ratio', 'Sales_TA']
    var_display = ['Intercept', 'Equity/TA', 'Quick Ratio', 'Sales/TA']

    row = 5
    for var, vd in zip(vars_order, var_display):
        ws.cell(row, 1, vd)
        ws.cell(row, 1).font = BOLD_FONT
        col = 2
        for dl in DATA_LABELS:
            key = f"{dl} - ROE"
            model = step1_results[key]
            coef = round(model.params[var], 4)
            pval = round(model.pvalues[var], 4)
            ws.cell(row, col, coef)
            ws.cell(row, col + 1, pval)
            ws.cell(row, col).font = sig_color(pval)
            col += 2
        style_data(ws, row, 9)
        row += 1

    row += 1
    for stat_name, stat_attr in [('R²', 'rsquared'), ('Adjusted R²', 'rsquared_adj'),
                                  ('F-statistic', 'fvalue'), ('F p-value', 'f_pvalue'),
                                  ('N', 'nobs')]:
        ws.cell(row, 1, stat_name)
        ws.cell(row, 1).font = BOLD_FONT
        col = 2
        for dl in DATA_LABELS:
            key = f"{dl} - ROE"
            model = step1_results[key]
            val = getattr(model, stat_attr)
            ws.cell(row, col, round(float(val), 4) if stat_attr != 'nobs' else int(val))
            col += 2
        style_data(ws, row, 9)
        row += 1

    row += 1
    ws.cell(row, 1,
            'Significance: Red = p<0.01, Blue = p<0.05, Green = p<0.10').font = Font(
        italic=True, name='Arial', size=9
    )

    ws.column_dimensions['A'].width = 15
    for c in 'BCDEFGHI':
        ws.column_dimensions[c].width = 14


def generate_table_4_4(wb, step1_results):
    """Table 4.4: Summary of All Regression Results."""
    ws = wb.create_sheet('Table 4.4')
    ws.cell(1, 1, 'Table 4.4: Summary of All Regression Results').font = TITLE_FONT

    headers = ['Model', 'Dep. Var.', 'Sample', 'R²', 'Adj. R²',
               'Equity/TA Coeff.', 'Equity/TA Sig.',
               'Quick Ratio Coeff.', 'Quick Ratio Sig.',
               'Sales/TA Coeff.', 'Sales/TA Sig.']
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, 11)

    row = 4
    model_num = 1
    for dep in DEP_VARS:
        for dl in DATA_LABELS:
            key = f"{dl} - {dep}"
            model = step1_results[key]
            ws.cell(row, 1, model_num)
            ws.cell(row, 2, dep)
            ws.cell(row, 3, dl)
            ws.cell(row, 4, round(model.rsquared, 4))
            ws.cell(row, 5, round(model.rsquared_adj, 4))
            ws.cell(row, 6, round(model.params['Equity_TA'], 4))
            ws.cell(row, 7, sig_stars(model.pvalues['Equity_TA']))
            ws.cell(row, 8, round(model.params['Quick_Ratio'], 4))
            ws.cell(row, 9, sig_stars(model.pvalues['Quick_Ratio']))
            ws.cell(row, 10, round(model.params['Sales_TA'], 4))
            ws.cell(row, 11, sig_stars(model.pvalues['Sales_TA']))
            style_data(ws, row, 11)
            row += 1
            model_num += 1

    row += 1
    ws.cell(row, 1, '*** p<0.01, ** p<0.05, * p<0.10').font = Font(
        italic=True, name='Arial', size=9
    )

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    for c in 'DEFGHIJK':
        ws.column_dimensions[c].width = 16


def generate_table_5_1(wb, step1_results, hyp_results):
    """Table 5.1: Hypotheses – Supported or Not Supported."""
    ws = wb.create_sheet('Table 5.1')
    ws.cell(1, 1, 'Table 5.1: Hypotheses – Supported or Not Supported').font = TITLE_FONT

    headers = ['Hypothesis', 'Dep. Var.', 'Indep. Var.', 'Expected Sign',
               'Germany 2021', 'Germany 2024', 'UK 2021', 'UK 2024']
    for i, h in enumerate(headers, 1):
        ws.cell(3, i, h)
    style_header(ws, 3, 8)

    row = 4
    for hyp in hyp_results:
        ws.cell(row, 1, hyp['Hypothesis'])
        ws.cell(row, 2, hyp['Dep'])
        ws.cell(row, 3, hyp['Variable'])
        ws.cell(row, 4, hyp['Expected'])
        for col_idx, dl in enumerate(DATA_LABELS, 5):
            outcome = hyp[dl]
            ws.cell(row, col_idx, outcome)
            if outcome == 'Supported':
                ws.cell(row, col_idx).fill = GREEN_FILL
            elif outcome == 'Rejected':
                ws.cell(row, col_idx).fill = RED_FILL
            else:
                ws.cell(row, col_idx).fill = GREY_FILL
        style_data(ws, row, 8)
        row += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    for c in 'EFGH':
        ws.column_dimensions[c].width = 18


# ============================================
# MAIN
# ============================================
def main():
    print("=" * 70)
    print("  Thesis Table Generator — Full Analysis Pipeline")
    print("=" * 70)

    # Step 1: Load data
    print("\n[STEP 1] Loading and cleaning data...")
    datasets = load_all_data()

    # Step 2: Descriptive statistics
    print("\n[STEP 2] Calculating descriptive statistics...")
    desc_stats = calculate_descriptive_stats(datasets)

    # Step 3: Correlation analysis
    print("\n[STEP 3] Calculating correlation matrices...")
    corr_data = calculate_correlations(datasets)

    # Step 4: OLS regressions (two-step)
    print("\n[STEP 4] Running OLS regressions (two-step approach)...")
    step1_results, step2_results = run_regressions(datasets)

    # Step 5: Hypothesis testing
    print("\n[STEP 5] Evaluating hypotheses...")
    hyp_results = evaluate_hypotheses(step1_results)

    # Step 6: Generate Excel tables
    print("\n[STEP 6] Generating Excel tables...")
    wb = Workbook()

    generate_table_2_1(wb)
    print("  ✓ Table 2.1: Summary of Key Empirical Studies")

    generate_table_2_2(wb)
    print("  ✓ Table 2.2: Summary of Hypotheses and Expected Signs")

    generate_table_3_1(wb, datasets)
    print("  ✓ Table 3.1: Sample Companies (Germany and UK)")

    generate_table_3_2(wb)
    print("  ✓ Table 3.2: Variable Definitions and Formulas")

    generate_table_3_3(wb, desc_stats)
    print("  ✓ Table 3.3: Descriptive Statistics (All Countries and Years)")

    generate_table_4_1(wb, corr_data)
    print("  ✓ Table 4.1: Correlation Matrices (Germany and UK)")

    generate_table_4_2(wb, step1_results)
    print("  ✓ Table 4.2: Regression Results – ROA (All 4 Models)")

    generate_table_4_3(wb, step1_results)
    print("  ✓ Table 4.3: Regression Results – ROE (All 4 Models)")

    generate_table_4_4(wb, step1_results)
    print("  ✓ Table 4.4: Summary of All Regression Results")

    generate_table_5_1(wb, step1_results, hyp_results)
    print("  ✓ Table 5.1: Hypotheses – Supported or Not Supported")

    # Save
    output_path = os.path.join(OUTPUT_DIR, 'All_Tables.xlsx')
    wb.save(output_path)

    print("\n" + "=" * 70)
    print(f"  All 10 tables saved to: {output_path}")
    print("=" * 70)

    # Print full regression summary to console
    print("\n\n" + "=" * 70)
    print("  FULL RESULTS SUMMARY")
    print("=" * 70)
    print(f"\n  {'Model':<25s} {'R²':>8s} {'Adj R²':>8s} {'Significant Variables'}")
    print("  " + "-" * 65)
    for key, model in step1_results.items():
        sig = [v for v in INDEP_VARS if model.pvalues[v] <= 0.10]
        print(f"  {key:<25s} {model.rsquared:>8.4f} {model.rsquared_adj:>8.4f} "
              f"{sig if sig else 'None'}")

    print(f"\n  Hypotheses: {sum(1 for r in hyp_results for l in DATA_LABELS if r[l] == 'Supported')} Supported, "
          f"{sum(1 for r in hyp_results for l in DATA_LABELS if r[l] == 'Rejected')} Rejected, "
          f"{sum(1 for r in hyp_results for l in DATA_LABELS if r[l] == 'Not Significant')} Not Significant")


if __name__ == '__main__':
    main()
